# Load tables from an Excel file, and write to a markdown file. Filter out rows and columns, based on defined lists
# Input 1: File name of Excel file (full path)
# Input 2: File name of markdown file (full path)

import sys
import openpyxl as pyxl


omit_columns = ["*"]         # Omit columns that includes any of these strings in the header
keep_rows = ["TENANT"]       # If column header is "ROLE", keep only rows that includes any of these strings in the row cells
included_tabs = ["Ericsson Private 5G","SDRs","Adv Routers","Worker Nodes"]       # List only tabs that includes any of these strings in the row cells



# Takes an Excel sheet as input and returns a list of tables in standard Python format
def get_tables(sheet):

    # Get list of rows in the Excel sheet
    row_list = []
    for tablerow in sheet.iter_rows(min_col=None, max_col=None, min_row=None, max_row=sheet.max_row, values_only=True):
        cell_list = []
        for cell in tablerow:
            cell_list.append(cell)
        row_list.append(cell_list)

    # Trim rightmost empty cells and top empty rows
    trimmed_row_list = []
    found_first_header = False
    for tablerow in row_list:
        while (len(tablerow) > 0) and (tablerow[-1] is None):
            tablerow.pop()
        if len(tablerow) > 0:
            found_first_header = True
        if found_first_header:
            trimmed_row_list.append(tablerow)

    # Split up into individual tables
    table_list = []
    prev_tablerow = ["dummy"]
    table = []
    for tablerow in trimmed_row_list:
        if ((len(tablerow) > 0) and (len(prev_tablerow) == 0)) or ((len(tablerow) == 0) and (len(prev_tablerow) > 0)):
            table_list.append(table)
            table = []
        table.append(tablerow)
        prev_tablerow = tablerow

    table_list.append(table)

    return table_list


# Takes a single table as input and returns a table where rows and columns have been filtered
def filter_rows_cols(table):
    headerrow = table[0]

    row_filtered_table = []
    for row, tablerow in enumerate(table):
        keeprow = True
        if row > 0:
            for col, cell in enumerate(tablerow):
                if "ROLE" in headerrow[col]:
                    if str(cell) in keep_rows:             # If column heading is "ROLE", then include only rows with keep_rows content in the row
                        keeprow = True  
                    else:
                        keeprow = False  

        if keeprow:
              row_filtered_table.append(tablerow)
    
    filtered_table = []
    for tablerow in row_filtered_table:
        newrow = []
        for col, cell in enumerate(tablerow):
            keepcolumn = True
            for omitstring in omit_columns:
                if omitstring in headerrow[col]:
                    keepcolumn = False

            if keepcolumn:                                 # Do not include columns with omit_columns content in the header
                newrow.append(cell)

        filtered_table.append(newrow)

    return filtered_table


# Takes a single table as input and returns a list of text lines with markdown
def get_table_md(table):
    headerrow = table[0]
    newlines = []

    # If the table is a "blank lines" table, then create corresponding markdown
    if len(headerrow) == 0:
        for row in range(len(table)):
            if row == 0:
                newlines.append("")
            elif row < (len(table)-1):
                newlines.append("<br/><br/>")
            else:
                newlines.append("<br/><br/>")
                newlines.append("")
   
    # If this is a normal table (not blank lines), then cteate corresponding markdown
    else:
        maxlen = []
        for col, cell in enumerate(headerrow):    
            maxlen.append(0)

        for row, tablerow in enumerate(table):          # Calculate width (maxlength) for each column
            for col, cell in enumerate(tablerow):
                if row == 0:
                    maxlen.append(0)
                if len(str(cell)) > maxlen[col]:
                    maxlen[col] = len(str(cell))

        for row, tablerow in enumerate(table):

            line = "|"
            for col, cell in enumerate(tablerow):
                line = line + str(cell)
                for i in range(maxlen[col]-len(str(cell))):
                    line = line + " "
                line = line + "|"
            
            newlines.append(line)

            if row == 0:                                # First row is header, so it needs a special "dash" line under it
                dashline = ""
                for ch in line:
                    if (ch == "|"):
                        dashline = dashline + "|"
                    else:
                        dashline = dashline + "-"
                newlines.append(dashline)

    return newlines


# Takes an Excel sheet as input and returns corresponding markup text lines
def excel_to_md(sheet):
    sheetlines = []
    sheetlines.append("")
    sheetlines.append("<br/><br/>")
    sheetlines.append("")
    sheetlines.append("## " + sheet.title)

    table_list = get_tables(sheet)

    for table in table_list:
        if len(table[0])>0:
            if table[0][0] == 'ID':
                sheetlines.append("### " + str(table[1][0]))
        filtered_table = filter_rows_cols(table)
        newlines = get_table_md(filtered_table)
        sheetlines.extend(newlines)

    return sheetlines



def main():

    # infname = "A:/Documents/expeca/lab-inventory/ExPECA-HW-Discovery.xlsx"  # Input Excel file name, full path
    # outfname = "A:/Documents/expeca/lab-inventory/inventory.md"      # Output markdown file name, full path

    # Command line input parameters
    # It has to be 2 command line parameters
    argc = len(sys.argv)
    if argc != 3:
        exit(0)  
    infname = sys.argv[1]                             # Input Excel file name, full path
    outfname = sys.argv[2]                            # Output markdown file name, full path

    linelist = []
    linelist.append("# Hardware Discovery")
    linelist.append("")
    linelist.append("*{{ git_revision_date_localized }}*")
    linelist.append("")

    wb = pyxl.load_workbook(infname)
 
    for sheet in wb:
        if sheet.title in included_tabs:
            sheetlines = excel_to_md(sheet)
            linelist.extend(sheetlines)

    with open(outfname, "w") as f:
        for line in linelist:
            f.write(line + "\n")

    return

if __name__ == "__main__":
    main()


